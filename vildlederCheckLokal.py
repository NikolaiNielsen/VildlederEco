"""
Forfatter: Nikolai Nielsen (NikolaiNielsen@outlook.dk)
Et script til at hjælpe med at gøre Vildlederøkonomien lidt nemmere:
- Gennemgår alle arkene set i variablen "SHEETS", finder alle indtastninger, og
  laver en oversigt over, hvad hver person har udlagt, under hvilken kategori,
  og med hvilken betalingsmetode.

Fremtidige forhåbninger:
- Automatisk genkendelse af, hvor de relevante data er, i hvert ark, så det er
  mere robust
- Grafisk brugerflade for dataindtastning
- Grafisk brugerflade til valg af fil

BRUG AF DETTE PROGRAM
- Programmet bruger pakker fra standardbiblioteket, samt "openpyxl" (version
  3.0.0), der skal installeres separat, for eksempel gennem pip. Den er også
  inkluderet med Anaconda distributionen.
- For at benytte dette program skal du have en lokal kopi af regnskabet i
  xlsx-format. Du skal ændre "WORKBOOK_NAME" til den relative sti til
  regnearket. Hvis regnearket er i samme mappe som dette program, kan du bare
  sætte WORKBOOK_NAME til at være filens navn (husk fil-typen!).
- Du skal ændre "SHEETS" til en liste af strings over hvilke ark, der skal
  tjekkes igennem. Som regel behøver denne ikke at blive ændret fra de normale,
  med mindre, der bliver ændret kategorier af øko-gruppen efter dette program
  er oprettet.
- Programmet forventer at data er indtastet fra kolonne A til H, og fra række 4
  og nedad. Det forventes af kolonnerne er som følger:
  - A: Bilagskode - kategorien af indkøbet. eksempelvis "Tema"
  - B: Bilagsnummer - Indkøbets nummer. Starter med 01, så 02 og så videre.
                      Sammen med bilagskoden udgør dette en samlet "ID" for
                      købet, eksempelvis Tema01.
  - C: NBI Rekvireringsnummer - som regel ikke vigtig. Bruges ikke her
  - D: Tekst - beskrivelse af købet
  - E: Status på bilag - Hvor ligger kvitteringen henne? (den skal helst i
                         mappen eller digitalt på dropbox eller lign.)
  - F: Beløb - hvor meget har dette indkøb kostet?
  - G: metode - hvordan indkøbet er foretaget. Der er 3 måder pt, set under
                "PAYMENT_METHODS". I arket forventes det at dette er et heltal,
                svarende til positionen i PAYMENT_METHODS (så 0=kontokort, etc)
  - H: Navn - hvem har foretaget købet? Der skelnes mellem store og små
              bogstaver!
- Det er egentlig kun kolonnerne A, B, F, G og H, der bliver brugt.
- "RANGES" bør kun ændres, hvis layouttet af regnearket ændres. Ligeledes skal
  "NUM_COLS" svare til hvor mange kolonner der skal læses fra (8, i dette
  tilfælde, da der skal læses fra A til H).
- Når disse ting er sørget for, skal programmet bare køres. Så bliver der
  automatisk indlæst og oprettet et nyt ark, "Opsummering", hvor der står en
  opsummering over, hvad hvert "navn" har købt, hvilke kvitteringer samt samlet
  beløb for hver betallingsmulighed. Så kan man nemt se, hvis man har skrevet
  et navn forkert, og hvor mange penge, hver person skal have tilbage gennem
  REJS-ud, samt hvor mange penge, der skal tilbagebetales til vejlederkontoen
  (i form af tilskud)
"""

import openpyxl
import pprint
from decimal import Decimal
from operator import itemgetter

WORKBOOK_NAME = "ProperVildleder2019.xlsx"
SHEETS = ['Tema', 'Mad']
PAYMENT_METHODS = ["kontokort", "vejlederkort", "personlig"]

RANGES = 'A4:H'
NUM_COLS = 8


def propagate_down(values, columnID=0):
    """
    Propagates the value of a given sublist down through empty sublists, in a
    given column. Ie, if the first entry of sublist 1 is "Mad", and the first
    entry of sublist 2 is None, then this code propagates the "Mad" down
    through the sublists, until it hits a non-None. Thus making sure that the
    column is fully populated.

    Assumes values is a list of lists.
    """
    for i in range(len(values)-1):
        if values[i+1][columnID] is None:
            values[i+1][columnID] = values[i][columnID]
    return values


def prepare_sheet_results(values, cols=[0, 1, 5, 6, 7]):
    """
    Takes the values of the sheet and prepares them for the summary.
    Assumes values is a list of lists.
    """

    # Values are returned as a list of lists, each of which contain the
    # cell contents. Default columns are
    # - Category
    # - receipt number
    # - price
    # - payment method
    # - name

    # Exclude empty rows - they have all None values.
    values = [x for x in values if x != [None]*NUM_COLS]

    # propagate down the category value
    values = propagate_down(values, 0)

    # Keep only certain columns.
    usable_values = [[row[i] for i in cols] for row in values]

    # Combine category and receipt number - pad with a single 0
    new_cat = [f'{row[0]}{int(row[1]):02d}' for row in usable_values]

    # include the new category in the listings (category, price, payment, name)
    combined = [[z[0], *z[1][2:]] for z in zip(new_cat, usable_values)]

    return combined


def get_values_from_sheet(sheet):
    # Find the max row count of the sheet and get the cells
    maxrow = sheet.max_row
    cells = sheet[f"{RANGES}{sheet.max_row}"]

    # Extract the raw values of each cell and store it in a list of lists
    values = [[i.internal_value for i in rows] for rows in cells]

    # prep the values and return the sheet
    prepped_values = prepare_sheet_results(values)
    return prepped_values


def combine_sheets(sheet_lists):
    """
    Combines the sheets. Assumes sheet_lists is a list of sheets (which
    themselves are a list of list)
    """
    combined = []
    for sheet in sheet_lists:
        combined = combined + sheet
    return combined


def process_data(sheet):
    """
    Takes the prepared data and processes it into the format used in the
    summary sheet.
    """
    # The vildleder has the structure of vildleder -> name -> payment method ->
    # receipt numbers and price
    vildleder = {}
    names = []
    # Get the names
    names = [row[-1] for row in sheet]
    unique_names = list(set(names))

    # Create the dictionary template
    for name in unique_names:
        vildleder[name] = {method: {"kvitteringer": [], "beløb": Decimal(0)}
                           for method in PAYMENT_METHODS}

    # populate the dictionary
    for row in sheet:
        receipt, price, method, name = row
        method = PAYMENT_METHODS[int(method)]
        # Round the amount to two decimals, and use Decimal. Probably an over
        # cautious approac.
        vildleder[name][method]['beløb'] += round(Decimal(price), 2)
        vildleder[name][method]['kvitteringer'].append(receipt)

    return vildleder


def populate_sheet(vildleder, workbook):
    """
    Takes the processed data from dictionary and populates the sheet.
    """
    sheet_name = "Opsummering"

    # Make sure the sheet is cleared
    if sheet_name in workbook:
        workbook.remove(workbook[sheet_name])

    workbook.create_sheet(sheet_name, 2)
    sheet = workbook[sheet_name]

    name_row = 4
    name_col = 2
    method_offset = 1
    price_offset = 2
    receipt_offset = 4
    names = sorted(list(vildleder.keys()))
    for n, name in enumerate(names):
        # Set names
        c = sheet.cell(row=name_row, column=name_col+n*len(PAYMENT_METHODS))
        c.value = name
        for m, method in enumerate(PAYMENT_METHODS):
            # Set payment methods
            c = sheet.cell(row=name_row+method_offset,
                           column=name_col+n*len(PAYMENT_METHODS)+m)
            c.value = method

            # Set total amount paid
            c = sheet.cell(row=name_row+price_offset,
                           column=name_col+n*len(PAYMENT_METHODS)+m)
            c.value = vildleder[name][method]["beløb"]

            # List receipts
            receipts = vildleder[name][method]["kvitteringer"]
            for i, receipt in enumerate(receipts):
                c = sheet.cell(row=name_row+receipt_offset+i,
                               column=name_col+n*len(PAYMENT_METHODS)+m)
                c.value = receipt

    # Set the "fucked receipts" title
    sheet[f'A{name_row}'] = "Kvitteringer Med Fejl"

    # Set column widths
    def as_text(val): return str(val) if val is not None else ""
    for column_cells in sheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(
                                column_cells[0].column)].width = length

    # Set title
    sheet['A1'] = sheet_name
    f = openpyxl.styles.Font(size=18)
    sheet['A1'].font = f
    sheet.merge_cells("A1:D1")


def main():
    """
    Open workbook. Get values from each worksheet in SHEETS, and combine them
    into one list. Process the data and create the summary sheet in the
    workbook.
    """
    wb = openpyxl.load_workbook(WORKBOOK_NAME)

    sheets = []
    for name in SHEETS:
        sheet = wb[name]
        values = get_values_from_sheet(sheet)
        sheets.append(values)

    sheet = combine_sheets(sheets)

    vildleder = process_data(sheet)

    populate_sheet(vildleder, wb)

    # Finally, save the workbook
    wb.save(WORKBOOK_NAME)


if __name__ == '__main__':
    main()
